from pathlib import Path
from typing import Dict, Any, List, Optional, Sequence
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def _autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_width, max(8, max_len + 2))


def save_summary_xlsx(
    path: Path,
    rows: List[Dict[str, Any]],
    meta: Dict[str, Any],
    raw_rows: Optional[Sequence[Sequence[Any]]] = None,
    raw_cycles: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """XLSX: Results (first) + Summary + Raw. No CSV required."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Results FIRST
    ws_res = wb.active
    ws_res.title = "Results"
    headers = [
        "nominal_voltage_V",
        "ref_channel",
        "ref_value_V",
        "dut_channel",
        "dut_value_mean_V",
        "lte_temp_mean_C",
        "dut_value_20C_V",
        "typeA_std_V",
        "typeA_u_mean_V",
        "u_ref_std_unc_V",
        "u_combined_V",
        "U_k2_V",
    ]
    ws_res.append(headers)
    for r in rows:
        ref = meta.get("ref", {})
        dut = meta.get("dut", {})
        ws_res.append([
            r.get("level_V"),
            ref.get("ch"),
            ref.get("value_v"),
            dut.get("ch"),
            r.get("u_dut_V"),
            r.get("t_mean_C"),
            r.get("u20_mean_V"),
            r.get("typeA_std_V"),
            r.get("typeA_u_mean_V"),
            r.get("u_ref_std_unc_V"),
            r.get("u_combined_V"),
            r.get("U_k2_V"),
        ])
    _autosize(ws_res)

    # Summary (extended)
    ws = wb.create_sheet("Summary")

    def _safe_cell(v: Any) -> Any:
        # openpyxl cannot write list/dict into a cell
        if isinstance(v, (list, tuple, dict, set)):
            try:
                return json.dumps(v, ensure_ascii=False)
            except Exception:
                return str(v)
        return v


    def _flatten(prefix: str, obj: Any, out: Dict[str, Any]) -> None:
        """Flatten nested dicts for the Summary sheet.

        Important: Excel cells can't store list/dict objects. We:
          - skip heavy raw payloads (raw_cycles/raw_rows) if they appear in meta
          - serialize lists/tuples/dicts to JSON strings (fallback)
        """
        SKIP_KEYS = {"raw_cycles", "raw_rows"}

        if isinstance(obj, dict):
            for k, v in obj.items():
                if k in SKIP_KEYS:
                    continue
                _flatten(f"{prefix}{k}.", v, out)
            return

        if isinstance(obj, (list, tuple)):
            out[prefix[:-1]] = json.dumps(obj, ensure_ascii=False)
            return

        # pydantic models or other objects -> stringify
        if isinstance(obj, (set,)):
            out[prefix[:-1]] = json.dumps(sorted(list(obj)), ensure_ascii=False)
            return

        out[prefix[:-1]] = obj

    flat_meta: Dict[str, Any] = {}
    _flatten("", meta, flat_meta)

    meta_cols = sorted(flat_meta.keys())
    row_cols = list(rows[0].keys()) if rows else []
    ws.append(meta_cols + row_cols)
    for r in rows or [{}]:
        ws.append([_safe_cell(flat_meta.get(k)) for k in meta_cols] + [_safe_cell(r.get(k)) for k in row_cols])
    _autosize(ws)

    # Raw sheet
    ws_raw = wb.create_sheet("Raw")

    def _write_raw_cycles(ws, cycles: List[Dict[str, Any]]) -> None:
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        row = 1
        for cyc in cycles:
            cyc_no = cyc.get("cycle")
            plus = cyc.get("plus") or []
            minus = cyc.get("minus") or []

            # Header row: merge A:D with Cycle N
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = ws.cell(row=row, column=1, value=f"Cycle {cyc_no}")
            c.font = bold
            c.alignment = center
            row += 1

            n = max(len(plus), len(minus))
            for i in range(n):
                p = plus[i] if i < len(plus) else None
                m = minus[i] if i < len(minus) else None

                ws.cell(row=row, column=1, value=(p.get("value_V") if p else None))
                ws.cell(row=row, column=2, value=(p.get("lte_temp_C") if p else None))
                ws.cell(row=row, column=3, value=(m.get("value_V") if m else None))
                ws.cell(row=row, column=4, value=(m.get("lte_temp_C") if m else None))
                row += 1

            # Blank line between cycles
            row += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14

    if raw_cycles:
        _write_raw_cycles(ws_raw, raw_cycles)
    elif raw_rows:
        # Backward compatible: append table rows as-is
        for row in raw_rows:
            ws_raw.append(list(row))
        _autosize(ws_raw, max_width=40)
    else:
        ws_raw.append(["t_epoch", "polarity", "value_V", "lte_temp_C", "env_t_C", "env_rh_pct", "env_p_kpa"])

    wb.save(str(path))
